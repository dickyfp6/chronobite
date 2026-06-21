import os
import sys
import openpyxl
from openpyxl.styles import Font

# Add paths to sys.path to access required modules dynamically based on location
current_dir = os.path.dirname(os.path.abspath(__file__))
if os.path.basename(current_dir) == 'Validasi':
    root_dir = os.path.abspath(os.path.join(current_dir, '..', '..'))
else:
    root_dir = os.path.abspath(os.path.join(current_dir, '..', '..', '..'))

sys.path.insert(0, root_dir)
sys.path.insert(0, os.path.join(root_dir, 'C. System Flow'))
sys.path.insert(0, os.path.join(root_dir, 'D. Model'))
sys.path.insert(0, os.path.join(root_dir, 'D. Model', 'greedy'))
sys.path.insert(0, os.path.join(root_dir, 'D. Model', 'Greedy Algorithm'))

from b_nutrition_service import NutritionService  # type: ignore
from greedy_interface import GreedyAlgorithmInterface  # type: ignore

USER1_BASE = {'gender': 'M', 'age': 28, 'weight': 68, 'height': 178, 'activity': 'moderate'}
USER2_BASE = {'gender': 'F', 'age': 55, 'weight': 83, 'height': 162, 'activity': 'light'}
USER3_BASE = {'gender': 'M', 'age': 34, 'weight': 51, 'height': 178, 'activity': 'vigorous'}

CASES = [
    # USER 1 — Normal + Single Disease (6 cases)
    {**USER1_BASE, 'sheet': 'U1_Normal',       'disease': ['normal'],                          'user_label': 'User 1'},
    {**USER1_BASE, 'sheet': 'U1_DM2',          'disease': ['dm2'],                             'user_label': 'User 1'},
    {**USER1_BASE, 'sheet': 'U1_Hypertension', 'disease': ['hypertension'],                    'user_label': 'User 1'},
    {**USER1_BASE, 'sheet': 'U1_CVD',          'disease': ['cvd'],                             'user_label': 'User 1'},
    {**USER1_BASE, 'sheet': 'U1_Cholesterol',  'disease': ['cholesterol'],                     'user_label': 'User 1'},
    {**USER1_BASE, 'sheet': 'U1_CKD',          'disease': ['ckd'],                             'user_label': 'User 1'},

    # USER 2 — Dual Disease (4 cases)
    {**USER2_BASE, 'sheet': 'U2_DM2+HT',       'disease': ['dm2', 'hypertension'],             'user_label': 'User 2'},
    {**USER2_BASE, 'sheet': 'U2_DM2+Chol',     'disease': ['dm2', 'cholesterol'],              'user_label': 'User 2'},
    {**USER2_BASE, 'sheet': 'U2_HT+CVD',       'disease': ['hypertension', 'cvd'],             'user_label': 'User 2'},
    {**USER2_BASE, 'sheet': 'U2_CKD+HT',       'disease': ['ckd', 'hypertension'],             'user_label': 'User 2'},

    # USER 3 — Triple Disease (3 cases)
    {**USER3_BASE, 'sheet': 'U3_DM2+HT+Chol',  'disease': ['dm2', 'hypertension', 'cholesterol'], 'user_label': 'User 3'},
    {**USER3_BASE, 'sheet': 'U3_CKD+DM2+HT',   'disease': ['ckd', 'dm2', 'hypertension'],         'user_label': 'User 3'},
    {**USER3_BASE, 'sheet': 'U3_HT+Chol+CVD',  'disease': ['hypertension', 'cholesterol', 'cvd'], 'user_label': 'User 3'},
]

ACTIVITY_MAP = {
    'moderate': 1.845,
    'light':    1.545,
    'vigorous': 2.2,
}

NUTRIENT_DISPLAY = {
    'energy_kcal': 'Energi (kcal)',
    'protein_g': 'Protein (g)',
    'fat_g': 'Lemak / Fat (g)',
    'carbohydrate_g': 'Karbohidrat (g)',
    'sodium_mg': 'Sodium (mg)',
    'potassium_mg': 'Potassium (mg)',
    'phosphorus_mg': 'Phosphorus (mg)',
    'cholesterol_mg': 'Cholesterol (mg)',
    'fiber_g': 'Serat / Fiber (g)',
    'calcium_mg': 'Calcium (mg)',
    'iron_mg': 'Iron / Zat Besi (mg)',
    'sugar_g': 'Sugar (g)',
    'saturated_fat_g': 'Saturated Fat (g)',
    'trans_fat_g': 'Trans Fat (g)',
}


SOFT_ORDER = [
    'vitamin_a_rae_mg',        # display: Vitamin A (mg RAE)
    'vitamin_c_mg',            # display: Vitamin C (mg)
    'vitamin_d_mg',            # display: Vitamin D (mg)
    'vitamin_e_mg',            # display: Vitamin E (mg)
    'vitamin_k_mg',            # display: Vitamin K (mg)
    'vitamin_b1_thiamin_mg',   # display: Vitamin B1 (Thiamine) (mg)
    'vitamin_b2_riboflavin_mg',# display: Vitamin B2 (Riboflavin) (mg)
    'vitamin_b3_niacin_mg',    # display: Vitamin B3 (Niacin) (mg)
    'vitamin_b5_pantothenic_acid_mg', # display: Vitamin B5 (Pantothenic Acid) (mg)
    'vitamin_b6_mg',           # display: Vitamin B6 (mg)
    'folate_mg',               # display: Vitamin B9 (Folate) (mg)
    'vitamin_b12_mg',          # display: Vitamin B12 (mg)
    'calcium_mg',              # display: Calcium (mg)
    'iron_mg',                 # display: Iron / Zat Besi (mg)
    'magnesium_mg',            # display: Magnesium (mg)
    'phosphorus_mg',           # display: Phosphorus (mg)
    'potassium_mg',            # display: Potassium (mg)
    'sodium_mg',               # display: Sodium (mg)
    'zinc_mg',                 # display: Zinc (mg)
    'copper_mg',               # display: Copper (mg)
    'manganese_mg',            # display: Manganese (mg)
    'selenium_mg',             # display: Selenium (mg)
    'water_g',                 # display: Air / Water (ml)
    'fiber_g',                 # display: Serat / Fiber (g)
]

SOFT_DISPLAY = {
    'vitamin_a_rae_mg': 'Vitamin A (mg RAE)',
    'vitamin_c_mg': 'Vitamin C (mg)',
    'vitamin_d_mg': 'Vitamin D (mg)',
    'vitamin_e_mg': 'Vitamin E (mg)',
    'vitamin_k_mg': 'Vitamin K (mg)',
    'vitamin_b1_thiamin_mg': 'Vitamin B1 (Thiamine) (mg)',
    'vitamin_b2_riboflavin_mg': 'Vitamin B2 (Riboflavin) (mg)',
    'vitamin_b3_niacin_mg': 'Vitamin B3 (Niacin) (mg)',
    'vitamin_b5_pantothenic_acid_mg': 'Vitamin B5 (Pantothenic Acid) (mg)',
    'vitamin_b6_mg': 'Vitamin B6 (mg)',
    'folate_mg': 'Vitamin B9 (Folate) (mg)',
    'vitamin_b12_mg': 'Vitamin B12 (mg)',
    'calcium_mg': 'Calcium (mg)',
    'iron_mg': 'Iron / Zat Besi (mg)',
    'magnesium_mg': 'Magnesium (mg)',
    'phosphorus_mg': 'Phosphorus (mg)',
    'potassium_mg': 'Potassium (mg)',
    'sodium_mg': 'Sodium (mg)',
    'zinc_mg': 'Zinc (mg)',
    'copper_mg': 'Copper (mg)',
    'manganese_mg': 'Manganese (mg)',
    'selenium_mg': 'Selenium (mg)',
    'water_g': 'Air / Water (ml)',
    'fiber_g': 'Serat / Fiber (g)',
}

DISEASE_MAP = {
    'normal': 'Normal',
    'dm2': 'DM2',
    'hypertension': 'Hipertensi',
    'cvd': 'CVD',
    'cholesterol': 'Kolesterol',
    'ckd': 'CKD'
}

def format_target(nutrient_key, min_val, max_val, tipe='range'):
    def fmt(val):
        if val is None or val == float('inf'):
            return ""
        if val == 0:
            return "0.0"
        if 'b12' in nutrient_key.lower():
            return f"{val:.6f}"
        elif val < 0.1:
            return f"{val:.4f}"
        else:
            return f"{val:.1f}"

    if tipe == 'max' and max_val is not None and max_val != float('inf'):
        return f"≤ {fmt(max_val)}"

    if min_val is not None and max_val is not None and max_val != float('inf'):
        return f"{fmt(min_val)} – {fmt(max_val)}"
    elif min_val is not None:
        return fmt(min_val)
    elif max_val is not None and max_val != float('inf'):
        return f"≤ {fmt(max_val)}"
    else:
        return "-"

def format_actual(nutrient_key, val):
    if val is None:
        return "-"
    if val == 0:
        return "0.0"
    if 'b12' in nutrient_key.lower():
        return f"{val:.6f}"
    elif val < 0.1:
        return f"{val:.4f}"
    else:
        return f"{val:.1f}"

def get_fulfillment(actual, min_val, max_val):
    if min_val is None:
        min_val = 0.0
    if max_val is None:
        max_val = float('inf')
    
    if actual < min_val:
        if min_val > 0:
            val = (actual / min_val) * 100
        else:
            val = 100.0
    elif actual > max_val:
        if max_val > 0 and max_val != float('inf'):
            val = (max_val / actual) * 100
        else:
            val = 100.0
    else:
        val = 100.0
    return f"{val:.1f}%"

def format_portion(p_val, is_drink=False):
    suffix = "ml" if is_drink else "g"
    if p_val == int(p_val):
        return f"{int(p_val)}{suffix}"
    return f"{p_val:.1f}{suffix}"

def main():
    try:
        nutrition_service = NutritionService()
    except Exception as e:
        print(f"[ERROR] Failed to initialize NutritionService: {e}")
        return

    wb = openpyxl.Workbook()
    # remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)  # type: ignore

    for i, case_raw in enumerate(CASES):
        case: Any = case_raw
        print(f"[{i+1}/13] Generating: {case['sheet']} ({case['disease']})...")
        ws = wb.create_sheet(title=str(case['sheet']))
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        try:
            profile = {
                'gender': case['gender'],
                'age': case['age'],
                'weight': case['weight'],
                'height': case['height'],
                'activity_factor': ACTIVITY_MAP[case['activity']],
                'disease': case['disease'],
                'food_preferences': []
            }
            
            analysis_result: Any = nutrition_service.calculate_nutrition_needs(profile)
            if not analysis_result.get('success'):
                raise ValueError(f"Nutrition analysis failed: {analysis_result.get('error')}")
                
            tdee = analysis_result['energy']['tdee']
            guidelines: Any = analysis_result['guidelines']
            food_database = analysis_result['food_data']['dataframe']

            
            greedy_engine = GreedyAlgorithmInterface(food_database, guidelines)
            menu_plan = greedy_engine.generate_menu_plan(profile, tdee)
            
            if not menu_plan:
                raise ValueError("Greedy menu generation returned None")
                
            # Extract courses and items
            b_main  = menu_plan.breakfast.courses['Main'].candidates[0]
            b_side  = menu_plan.breakfast.courses['Side'].candidates[0]
            b_drink = menu_plan.breakfast.courses['Drink'].candidates[0]

            l_main  = menu_plan.lunch.courses['Main'].candidates[0]
            l_side  = menu_plan.lunch.courses['Side'].candidates[0]
            l_drink = menu_plan.lunch.courses['Drink'].candidates[0]

            d_main  = menu_plan.dinner.courses['Main'].candidates[0]
            d_side  = menu_plan.dinner.courses['Side'].candidates[0]
            d_drink = menu_plan.dinner.courses['Drink'].candidates[0]

            snack   = menu_plan.snack.candidates[0]

            # Calories per meal
            breakfast_cal = b_main.energy_kcal + b_side.energy_kcal + b_drink.energy_kcal
            lunch_cal     = l_main.energy_kcal + l_side.energy_kcal + l_drink.energy_kcal
            dinner_cal    = d_main.energy_kcal + d_side.energy_kcal + d_drink.energy_kcal
            snack_cal     = snack.energy_kcal
            
            # --- SECTION 1: IDENTITAS USER ---
            ws.merge_cells('A1:B1')
            ws['A1'].value = "IDENTITAS USER"
            ws['A1'].font = Font(bold=True)

            
            gender_display = 'Pria' if case['gender'] == 'M' else 'Wanita'
            activity_display = {
                'moderate': 'Sedang',
                'light':    'Rendah',
                'vigorous': 'Tinggi',
            }[case['activity']]
            disease_display = " + ".join([DISEASE_MAP.get(d, d.title()) for d in case['disease']])
            
            ws['A2'] = "Usia"
            ws['B2'] = f"{case['age']} tahun"
            ws['A3'] = "Jenis Kelamin"
            ws['B3'] = gender_display
            ws['A4'] = "Berat Badan"
            ws['B4'] = f"{case['weight']} kg"
            ws['A5'] = "Tinggi Badan"
            ws['B5'] = f"{case['height']} cm"
            ws['A6'] = "Tingkat Aktivitas"
            ws['B6'] = activity_display
            ws['A7'] = "Kondisi Penyakit"
            ws['B7'] = disease_display
            
            # --- SECTION 2: KEBUTUHAN NUTRISI USER ---
            ws['A9'].value = "KEBUTUHAN NUTRISI USER"
            ws['A9'].font = Font(bold=True)

            
            bmr = analysis_result['energy']['bmr']
            tdee = analysis_result['energy']['tdee']
            
            ws['A10'] = "BMR (kcal)"
            ws['B10'] = round(bmr, 1)
            ws['A11'] = "TDEE (kcal)"
            ws['B11'] = round(tdee, 1)
            
            # Guidelines ranges helper
            def format_min_max(nutrient_key):
                nut = guidelines['nutrients'].get(nutrient_key, {})
                min_val = nut.get('min')
                max_val = nut.get('max')
                tipe_val = nut.get('tipe', 'range')
                if min_val is None and max_val is None:
                    if nutrient_key == 'energy_kcal':
                        min_val = tdee * 0.95
                        max_val = tdee * 1.05
                    else:
                        return '-'
                if tipe_val == 'max' and max_val is not None and max_val != float('inf'):
                    return f"≤ {max_val:.1f}"
                if max_val == float('inf') or max_val is None:
                    return f"≥ {min_val:.1f}" if min_val is not None else "-"
                if min_val is None or min_val == 0:
                    return f"≤ {max_val:.1f}"
                return f"{min_val:.1f} – {max_val:.1f}"

                
            ws['A12'] = "Energi Harian (kcal)"
            ws['B12'] = format_min_max('energy_kcal')
            ws['A13'] = "Protein (g)"
            ws['B13'] = format_min_max('protein_g')
            ws['A14'] = "Lemak / Fat (g)"
            ws['B14'] = format_min_max('fat_g')
            ws['A15'] = "Karbohidrat (g)"
            ws['B15'] = format_min_max('carbohydrate_g')
            
            # Water
            ws['A16'] = "Air / Water (ml)"
            water_nut = guidelines['nutrients'].get('water_g')
            if water_nut:
                w_min = water_nut.get('min')
                w_max = water_nut.get('max')
                if w_min is not None and w_max is not None and w_max != float('inf'):
                    water_str = f"{w_min:.1f} – {w_max:.1f}"
                elif w_min is not None:
                    water_str = f"≥ {w_min:.1f}"
                else:
                    water_str = '-'
            else:
                water_str = '-'
            ws['B16'] = water_str
            
            # --- SECTION 3: MENU MAKANAN ---
            ws['A18'] = "Waktu"
            ws['B18'] = "Makan"
            ws['C18'] = "Menu / Makanan"
            ws['D18'] = "Porsi (g/ml)"
            ws['E18'] = "Kalori (kcal)"
            for col in ['A18', 'B18', 'C18', 'D18', 'E18']:
                ws[col].font = Font(bold=True)
                
            menu_data = [
                ("Breakfast", "Main Course", b_main.food_name, format_portion(b_main.portion_gram), b_main.energy_kcal),
                ("Breakfast", "Side Dish", b_side.food_name, format_portion(b_side.portion_gram), b_side.energy_kcal),
                ("Breakfast", "Drink", b_drink.food_name, format_portion(b_drink.portion_gram, is_drink=True), b_drink.energy_kcal),
                (None, None, None, None, None), # blank row
                ("Lunch", "Main Course", l_main.food_name, format_portion(l_main.portion_gram), l_main.energy_kcal),
                ("Lunch", "Side Dish", l_side.food_name, format_portion(l_side.portion_gram), l_side.energy_kcal),
                ("Lunch", "Drink", l_drink.food_name, format_portion(l_drink.portion_gram, is_drink=True), l_drink.energy_kcal),
                (None, None, None, None, None), # blank row
                ("Dinner", "Main Course", d_main.food_name, format_portion(d_main.portion_gram), d_main.energy_kcal),
                ("Dinner", "Side Dish", d_side.food_name, format_portion(d_side.portion_gram), d_side.energy_kcal),
                ("Dinner", "Drink", d_drink.food_name, format_portion(d_drink.portion_gram, is_drink=True), d_drink.energy_kcal),
                (None, None, None, None, None), # blank row
                ("Snack", "Snack", snack.food_name, format_portion(snack.portion_gram), snack.energy_kcal),
                (None, None, None, None, None), # blank row
            ]
            
            current_row = 19
            for row_data in menu_data:
                if row_data[0] is not None:
                    ws.cell(row=current_row, column=1, value=row_data[0])
                    ws.cell(row=current_row, column=2, value=row_data[1])
                    ws.cell(row=current_row, column=3, value=row_data[2])
                    ws.cell(row=current_row, column=4, value=row_data[3])
                    ws.cell(row=current_row, column=5, value=round(row_data[4], 1))
                current_row += 1
                
            # --- SECTION 4: DISTRIBUSI KALORI ---
            ws.cell(row=current_row, column=1, value="Waktu Makan").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Kalori (kcal)").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value="% Target").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value="Keterangan").font = Font(bold=True)
            for col_idx in range(1, 5):
                ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            current_row += 1
            
            dist_data = [
                ("Breakfast (Sarapan)", round(breakfast_cal, 1), "23.75%", "Konstanta sistem"),
                ("Lunch (Makan Siang)", round(lunch_cal, 1), "33.75%", "Konstanta sistem"),
                ("Dinner (Makan Malam)", round(dinner_cal, 1), "28.75%", "Konstanta sistem"),
                ("Snack", round(snack_cal, 1), "13.75%", "Konstanta sistem"),
                ("TOTAL", round(breakfast_cal + lunch_cal + dinner_cal + snack_cal, 1), "100%", "")
            ]
            
            for row_data in dist_data:
                ws.cell(row=current_row, column=1, value=row_data[0])
                ws.cell(row=current_row, column=2, value=row_data[1])
                ws.cell(row=current_row, column=3, value=row_data[2])
                ws.cell(row=current_row, column=4, value=row_data[3])
                if row_data[0] == "TOTAL":
                    for col_idx in range(1, 5):
                        ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                current_row += 1
                
            current_row += 1 # blank row
            
            # --- SECTION 5: EVALUASI HARD CONSTRAINT ---
            ws.cell(row=current_row, column=1, value="Nutrisi").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Min – Max").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value="Aktual").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value="Keterpenuhan (%)").font = Font(bold=True)
            for col_idx in range(1, 5):
                ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            current_row += 1
            
            actual_nutrients = {
                'energy_kcal':    menu_plan.total_daily_calories,
                'protein_g':      menu_plan.total_daily_protein_g,
                'carbohydrate_g': menu_plan.total_daily_carb_g,
                'fat_g':          menu_plan.total_daily_fat_g,
            }
            if hasattr(menu_plan, 'daily_micronutrients') and menu_plan.daily_micronutrients:
                actual_nutrients.update(menu_plan.daily_micronutrients)
                
            # Define the fixed order and display names for the primary nutrients
            primary_hard_keys = [
                ('energy_kcal', 'Energy (kcal)'),
                ('protein_g', 'Protein (g)'),
                ('fat_g', 'Fat (g)'),
                ('carbohydrate_g', 'Carbohydrate (g)'),
                ('sodium_mg', 'Sodium (mg)'),
                ('sugar_g', 'Sugar (g)'),
                ('cholesterol_mg', 'Cholesterol (mg)'),
            ]

            
            # Print the primary nutrients in the exact fixed order
            printed_keys = set()

            for nutrient_key, display_name in primary_hard_keys:
                printed_keys.add(nutrient_key)
                constraint = guidelines['nutrients'].get(nutrient_key)
                
                min_v = None
                max_v = None
                tipe_val = 'range'
                if constraint:
                    min_v = constraint.get('min')
                    max_v = constraint.get('max')
                    tipe_val = constraint.get('tipe', 'range')
                elif nutrient_key == 'energy_kcal':
                    min_v = tdee * 0.95
                    max_v = tdee * 1.05
                    tipe_val = 'range'

                limit_str = format_target(nutrient_key, min_v, max_v, tipe=tipe_val)
                actual_val = actual_nutrients.get(nutrient_key, 0.0)
                actual_str = format_actual(nutrient_key, actual_val)
                
                if min_v is None and max_v is None:
                    fulfillment_str = "-"
                else:
                    fulfillment_str = get_fulfillment(actual_val, min_v, max_v)
                
                ws.cell(row=current_row, column=1, value=display_name)
                ws.cell(row=current_row, column=2, value=limit_str)
                ws.cell(row=current_row, column=3, value=actual_str)
                ws.cell(row=current_row, column=4, value=fulfillment_str)
                current_row += 1

            # Print any OTHER hard constraints that might be active for this profile (e.g. saturated_fat_g, potassium_mg, etc.)
            for nutrient_key, constraint in guidelines['nutrients'].items():
                if nutrient_key in printed_keys:
                    continue
                if constraint.get('hard_soft_type') == 'HARD' and constraint.get('constraint_type') != 'unlimited':
                    display_name = NUTRIENT_DISPLAY.get(nutrient_key, nutrient_key)
                    min_v = constraint.get('min')
                    max_v = constraint.get('max')
                    tipe_val = constraint.get('tipe', 'range')
                    
                    limit_str = format_target(nutrient_key, min_v, max_v, tipe=tipe_val)
                    actual_val = actual_nutrients.get(nutrient_key, 0.0)
                    actual_str = format_actual(nutrient_key, actual_val)
                    fulfillment_str = get_fulfillment(actual_val, min_v, max_v)
                    
                    ws.cell(row=current_row, column=1, value=display_name)
                    ws.cell(row=current_row, column=2, value=limit_str)
                    ws.cell(row=current_row, column=3, value=actual_str)
                    ws.cell(row=current_row, column=4, value=fulfillment_str)
                    current_row += 1
                    
            current_row += 1 # blank row
            
            # --- SECTION 6: EVALUASI SOFT CONSTRAINT ---
            ws.cell(row=current_row, column=1, value="Nutrisi").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="Target (DRI)").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value="Aktual").font = Font(bold=True)
            ws.cell(row=current_row, column=4, value="Keterpenuhan (%)").font = Font(bold=True)
            for col_idx in range(1, 5):
                ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            current_row += 1
            
            hard_nutrients = {
                k for k, v in guidelines['nutrients'].items()
                if v.get('hard_soft_type') == 'HARD' and v.get('constraint_type') != 'unlimited'
            }
            
            for nutrient_key in SOFT_ORDER:
                if nutrient_key in hard_nutrients:
                    continue
                constraint = guidelines['nutrients'].get(nutrient_key)
                if not constraint:
                    continue
                    
                display_name = SOFT_DISPLAY.get(nutrient_key, nutrient_key)
                min_v = constraint.get('min')
                max_v = constraint.get('max')
                tipe_val = constraint.get('tipe', 'range')
                
                limit_str = format_target(nutrient_key, min_v, max_v, tipe=tipe_val)
                actual_val = actual_nutrients.get(nutrient_key, 0.0)
                actual_str = format_actual(nutrient_key, actual_val)
                fulfillment_str = get_fulfillment(actual_val, min_v, max_v)
                
                ws.cell(row=current_row, column=1, value=display_name)
                ws.cell(row=current_row, column=2, value=limit_str)
                ws.cell(row=current_row, column=3, value=actual_str)
                ws.cell(row=current_row, column=4, value=fulfillment_str)
                current_row += 1
                
            print(f"  -> Done: {menu_plan.total_daily_calories:.0f} kcal")
            
        except Exception as e:
            print(f"  [ERROR] {e}")
            import traceback
            traceback.print_exc()
            
            # Clear worksheet
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
            
            # Write error message in A1
            ws['A1'].value = "GAGAL GENERATE MENU"
            ws['A1'].font = Font(bold=True)

    output_dir = os.path.join(current_dir, 'output')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'validasi_greedy.xlsx')
    wb.save(output_path)
    print(f"\n[SUCCESS] Saved validation results to: {output_path}")

if __name__ == "__main__":
    main()
