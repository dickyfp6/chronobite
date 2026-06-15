import openpyxl
import os
import sys

# Set stdout encoding to utf-8 to print Unicode characters safely
sys.stdout.reconfigure(encoding='utf-8')

path = r"c:\Users\USERR\Documents\0. Mata Kuliah\8 -TA\Code\TugasAkhirDSS\H. Validasi\output\validasi_greedy_temp.xlsx"
if os.path.exists(path):
    wb = openpyxl.load_workbook(path)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for r in range(1, ws.max_row + 1):
            val_a = ws.cell(row=r, column=1).value
            if val_a and "Sugar (g)" in str(val_a):
                target = ws.cell(row=r, column=2).value
                actual = ws.cell(row=r, column=3).value
                fulfillment = ws.cell(row=r, column=4).value
                print(f"Sheet: {sheetname} | Row {r} | A: {val_a} | B: {target} | C: {actual} | D: {fulfillment}")
else:
    print("File not found")
