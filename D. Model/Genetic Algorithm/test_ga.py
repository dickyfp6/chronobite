"""
TEST GENETIC ALGORITHM - Integration with NutritionService
==========================================================

Test file untuk GA dengan NutritionService
- Input user data
- Calculate nutrition needs via NutritionService
- Run GA
- Display hasil
"""

import sys
import os
import random
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add paths untuk import
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
system_flow_path = os.path.join(project_root, 'C. System Flow')
genetic_algorithm_path = os.path.dirname(os.path.abspath(__file__))

sys.path.insert(0, system_flow_path)
sys.path.insert(0, genetic_algorithm_path)

# ============ MODE SWITCH ============
# Set to True to use interactive input via CLI
# Set to False to use hardcoded default values
USE_INTERACTIVE_INPUT = True
# ====================================

# Import GA engine
from ga_v1 import (
    run_ga, display_solution, generate_meal_options, display_meal_options, 
    display_fitness_details, MEAL_INDICES, calculate_total_nutrition, 
    SLOT_NAMES, CHROMOSOME_SIZE, calculate_portion_sizes_dynamic, display_portion_summary_dynamic,
    local_search, display_nutrition_analysis_table
)

# Import NutritionService
try:
    from nutrition_service import NutritionService
    print("✓ NutritionService imported successfully")
except ImportError as e:
    print(f"✗ Cannot import NutritionService: {e}")
    sys.exit(1)

# Import input handler
try:
    from modules.input_handler import get_user_input
    print("✓ Input handler imported successfully")
except ImportError as e:
    print(f"✗ Cannot import input handler: {e}")
    sys.exit(1)


# ════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS: Percentage Fulfillment & Status Categorization
# ════════════════════════════════════════════════════════════════════════

def calculate_fulfillment_percentage(value, min_val, max_val):
    """
    Calculate nutrient fulfillment percentage based on range.
    
    Logic:
    - If value < min: percent = (value / min) * 100
    - Elif value > max: percent = (max / value) * 100
    - Else: percent = 100
    
    Args:
        value: Actual nutrient value
        min_val: Minimum target
        max_val: Maximum target
    
    Returns:
        float: Percentage (0-100+)
    """
    if min_val == 0 and max_val == float('inf'):
        # No constraint
        return 100.0
    
    if value < min_val:
        # Below minimum - calculate deficit percentage
        percent = (value / min_val * 100) if min_val > 0 else 0
    elif value > max_val:
        # Above maximum - calculate excess percentage
        percent = (max_val / value * 100) if value > 0 else 0
    else:
        # Within range - 100%
        percent = 100.0
    
    return percent


def get_status_category(percent):
    """
    Categorize nutrient fulfillment status based on percentage.
    
    Categories:
    - >= 95%: Excellent ✨
    - >= 85%: Good 🟢
    - >= 70%: Fair 🟡
    - < 70%: Poor 🔴
    
    Args:
        percent: Fulfillment percentage
    
    Returns:
        tuple: (status_text, emoji, category_color)
    """
    if percent >= 95:
        return ("Excellent", "✨", "green")
    elif percent >= 85:
        return ("Good", "🟢", "green")
    elif percent >= 70:
        return ("Fair", "🟡", "yellow")
    else:
        return ("Poor", "🔴", "red")


def format_fulfillment_display(value, min_val, max_val, unit):
    """
    Format nutrient value display dengan percentage.
    
    Example: 197 g / 241 g → 81.7% (Fair)
    
    Args:
        value: Actual value
        min_val: Minimum target
        max_val: Maximum target
        unit: Unit string
    
    Returns:
        tuple: (display_string, percentage, category)
    """
    percent = calculate_fulfillment_percentage(value, min_val, max_val)
    status_text, emoji, category = get_status_category(percent)
    
    # Format based on constraint type
    if min_val == 0 and max_val == float('inf'):
        display_str = f"{value:.1f} {unit}"
    elif min_val == max_val:
        # Target value (exact)
        display_str = f"{value:.1f} / {min_val:.1f} {unit}"
    else:
        # Range
        if max_val == float('inf'):
            display_str = f"{value:.1f} / min {min_val:.1f} {unit}"
        else:
            display_str = f"{value:.1f} ({min_val:.1f}-{max_val:.1f}) {unit}"
    
    return display_str, percent, status_text, emoji


def get_simple_user_input(interactive=False):
    """
    Get user input either interactively or use defaults
    
    Args:
        interactive: If True, prompt for input. If False, use defaults.
    
    Returns:
        dict: user_input untuk NutritionService
    """
    print("\n" + "="*70)
    print("MEAL PLANNING SYSTEM - USER INPUT")
    print("="*70)
    
    if interactive:
        print("\n(Press Enter untuk gunakan default values)")
        
        gender = input("Gender (M/F) [M]: ").strip() or "M"
        age = int(input("Age (18-100) [25]: ").strip() or "25")
        weight = float(input("Weight (kg) [70]: ").strip() or "70")
        height = float(input("Height (cm) [170]: ").strip() or "170")
        activity = input("Activity Factor (1.4-2.2) [1.55]: ").strip() or "1.55"
        activity_factor = float(activity)
        
        print("\nHealth Conditions (comma-separated):")
        print("  Valid: normal, dm2, hypertension, cvd, cholesterol, ckd")
        disease_input = input("Diseases [normal]: ").strip() or "normal"
        disease = [d.strip() for d in disease_input.split(",")]
        
        print("\nFood Preferences (comma-separated):")
        print("  Valid: Asian, Western, Mediterranean, Generic")
        preferences_input = input("Preferences [Asian, Western]: ").strip() or "Asian, Western"
        food_preferences = [p.strip() for p in preferences_input.split(",")]
    else:
        # Use defaults (non-interactive)
        # print("\n(Using default values)")  # Commented out - using real input
        gender = "M"
        age = 25
        weight = 70.0
        height = 170.0
        activity_factor = 1.55
        disease = ["normal"]
        food_preferences = ["Asian", "Western"]
    
    user_input = {
        'gender': gender,
        'age': age,
        'weight': weight,
        'height': height,
        'activity_factor': activity_factor,
        'disease': disease,
        'food_preferences': food_preferences
    }
    
    return user_input


# ════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL - STEP 11
# ════════════════════════════════════════════════════════════════════════

# Mapping untuk slot ke label waktu makan
SLOT_LABEL_MAP = {
    'breakfast_main':  'Breakfast — Main Course',
    'breakfast_side':  'Breakfast — Side Dish',
    'breakfast_drink': 'Breakfast — Drink',
    'lunch_main':      'Lunch — Main Course',
    'lunch_side':      'Lunch — Side Dish',
    'lunch_drink':     'Lunch — Drink',
    'dinner_main':     'Dinner — Main Course',
    'dinner_side':     'Dinner — Side Dish',
    'dinner_drink':    'Dinner — Drink',
    'snack':           'Snack',
}

# Mapping untuk keys di total_nutrition vs keys di guidelines
HARD_NUTRIENT_KEY_MAP = {
    'energy':        ['energy_kcal', 'calories', 'energy'],
    'protein':       ['protein_g', 'protein'],
    'fat':           ['fat_g', 'total_fat_g', 'fat'],
    'carbohydrate':  ['carbohydrate_g', 'carbs_g', 'carbohydrate'],
    'sodium':        ['sodium_mg', 'sodium'],
    'sugar':         ['sugar_g', 'total_sugar_g', 'sugars_g'],
    'cholesterol':   ['cholesterol_mg', 'cholesterol'],
}

SOFT_NUTRIENT_KEY_MAP = {
    'vitamin_a':        ['vitamin_a_rae_mcg', 'vitamin_a_mcg', 'vitamin_a'],
    'vitamin_c':        ['vitamin_c_mg', 'vitamin_c'],
    'vitamin_d':        ['vitamin_d_mcg', 'vitamin_d_iu', 'vitamin_d'],
    'vitamin_e':        ['vitamin_e_mg', 'vitamin_e'],
    'vitamin_k':        ['vitamin_k_mcg', 'vitamin_k'],
    'thiamine':         ['thiamine_mg', 'vitamin_b1_mg', 'thiamine'],
    'riboflavin':       ['riboflavin_mg', 'vitamin_b2_mg', 'riboflavin'],
    'niacin':           ['niacin_mg', 'niacin'],
    'pantothenic_acid': ['pantothenic_acid_mg', 'vitamin_b5_mg', 'pantothenic_acid'],
    'vitamin_b6':       ['vitamin_b6_mg', 'vitamin_b6'],
    'folate':           ['folate_mcg', 'vitamin_b9_mcg', 'folate_dfe_mcg', 'folate'],
    'vitamin_b12':      ['vitamin_b12_mcg', 'vitamin_b12'],
    'calcium':          ['calcium_mg', 'calcium'],
    'iron':             ['iron_mg', 'iron'],
    'magnesium':        ['magnesium_mg', 'magnesium'],
    'phosphorus':       ['phosphorus_mg', 'phosphorus'],
    'potassium':        ['potassium_mg', 'potassium'],
    'sodium':           ['sodium_mg', 'sodium'],
    'zinc':             ['zinc_mg', 'zinc'],
    'copper':           ['copper_mg', 'copper'],
    'manganese':        ['manganese_mg', 'manganese'],
    'selenium':         ['selenium_mcg', 'selenium'],
    'water':            ['water_g', 'water_ml', 'water'],
    'fiber':            ['fiber_g', 'dietary_fiber_g', 'total_fiber_g', 'fiber'],
}


def lookup_nutrition(total_nutrition, key, key_map):
    """
    Lookup nutrition value dari total_nutrition dict dengan fallback keys.
    
    Args:
        total_nutrition: Dict hasil calculate_total_nutrition
        key: Primary key untuk dicari
        key_map: Dict mapping primary key ke candidate keys
    
    Returns:
        float atau None: Nilai nutrisi atau None jika tidak ditemukan
    """
    candidates = key_map.get(key, [key])
    for candidate in candidates:
        if candidate in total_nutrition:
            return total_nutrition[candidate]
    return None


def find_column_in_df(df, priority_list):
    """
    Find first existing column dari priority list.
    
    Args:
        df: DataFrame untuk dicek
        priority_list: List kolom dalam urutan prioritas
    
    Returns:
        str atau None: Nama kolom pertama yang ada, atau None
    """
    for col in priority_list:
        if col in df.columns:
            return col
    return None


def export_to_excel(filename, user_input, nutrition_result, guidelines_all, 
                   selected_df, portion_result_df, guidelines, tdee):
    """
    Export meal planning results ke Excel dengan format template validasi ahli gizi.
    
    Args:
        filename: Nama file Excel (.xlsx)
        user_input: Dict input user
        nutrition_result: Result dari NutritionService.calculate_nutrition_needs()
        guidelines_all: Dict guidelines sebelum di-split (untuk Bagian 2)
        selected_df: DataFrame hasil pemilihan user (10 items)
        portion_result_df: DataFrame hasil portion sizing
        guidelines: Dict {'hard': {...}, 'soft': {...}} untuk Bagian 5 & 6
        tdee: TDEE user
    
    Returns:
        bool: True jika sukses, False jika gagal
    """
    try:
        print("\n" + "="*70)
        print("STEP 11: EXPORT HASIL KE EXCEL")
        print("="*70)
        
        # ════════════════════════════════════════════════════════════════════════
        # LANGKAH PERTAMA — INSPECT KOLOM
        # ════════════════════════════════════════════════════════════════════════
        print("\n[DEBUG export] Inspecting columns...")
        
        # Inspect total nutrition keys
        total_nutrition = calculate_total_nutrition(selected_df)
        print(f"[DEBUG export] Total nutrition keys: {list(total_nutrition.keys())}")
        
        # Inspect portion_result_df columns
        print(f"[DEBUG export] portion_result_df columns: {list(portion_result_df.columns)}")
        
        # ════════════════════════════════════════════════════════════════════════
        # SETUP WORKBOOK DAN SHEET
        # ════════════════════════════════════════════════════════════════════════
        wb = Workbook()
        ws = wb.active
        
        # Tentukan sheet name berdasarkan disease label
        disease_label = ", ".join(user_input['disease']).upper()
        sheet_name = f"Case - {disease_label}"[:31]  # Excel max 31 karakter
        ws.title = sheet_name
        
        # Set column widths
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 22
        
        # Define styling
        title_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        title_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        header_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # ════════════════════════════════════════════════════════════════════════
        # BAGIAN 1 — PROFIL USER
        # ════════════════════════════════════════════════════════════════════════
        section_1_row = current_row
        
        # Title Bagian 1
        ws[f'A{current_row}'] = "BAGIAN 1 — PROFIL USER"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        # Data Bagian 1
        profile_data = [
            ('Usia', str(user_input['age'])),
            ('Jenis Kelamin', user_input['gender']),
            ('Berat Badan (kg)', f"{user_input['weight']:.1f}"),
            ('Tinggi Badan (cm)', f"{user_input['height']:.1f}"),
            ('Tingkat Aktivitas', f"{user_input['activity_factor']:.2f}"),
            ('Kondisi Penyakit', ", ".join(user_input['disease'])),
        ]
        
        for label, value in profile_data:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
        
        # ════════════════════════════════════════════════════════════════════════
        # BAGIAN 2 — KEBUTUHAN NUTRISI USER (Berdampingan dengan Bagian 1)
        # ════════════════════════════════════════════════════════════════════════
        # Kembali ke row yang sama dengan Bagian 1
        section_2_row = section_1_row
        
        # Title Bagian 2
        ws[f'D{section_2_row}'] = "BAGIAN 2 — KEBUTUHAN NUTRISI USER"
        ws[f'D{section_2_row}'].font = title_font
        ws[f'D{section_2_row}'].fill = title_fill
        ws.merge_cells(f'D{section_2_row}:E{section_2_row}')
        section_2_row += 1
        
        # Helper function untuk safe get dari guidelines_all
        def safe_get_min(guidelines_all, key):
            return guidelines_all.get(key, {}).get('min', 0)
        
        # Data Bagian 2
        # TODO: ganti jika ada adjusted energy key di nutrition_result['energy']
        energy_daily = nutrition_result['energy'].get('daily_energy') or \
                       nutrition_result['energy'].get('adjusted_energy') or \
                       nutrition_result['energy'].get('target_energy') or \
                       nutrition_result['energy']['tdee']
        
        nutrition_data = [
            ('BMR (kcal)', f"{nutrition_result['energy']['bmr']:.1f}"),
            ('TDEE (kcal)', f"{nutrition_result['energy']['tdee']:.0f}"),
            ('Energi Harian (kcal)', f"{energy_daily:.0f}"),
            ('Protein (g)', f"{safe_get_min(guidelines_all, 'protein'):.1f}"),
            ('Lemak / Fat (g)', f"{safe_get_min(guidelines_all, 'fat'):.1f}"),
            ('Karbohidrat (g)', f"{safe_get_min(guidelines_all, 'carbohydrate'):.1f}"),
        ]
        
        # Tambahkan Water dengan fallback
        water_val = safe_get_min(guidelines_all, 'water')
        if water_val == 0:
            water_val = 2000  # Fallback
        nutrition_data.append(('Air / Water (ml)', f"{water_val:.0f}"))
        
        for label, value in nutrition_data:
            ws[f'D{section_2_row}'] = label
            ws[f'E{section_2_row}'] = value
            ws[f'D{section_2_row}'].border = border
            ws[f'E{section_2_row}'].border = border
            section_2_row += 1
        
        # Advance current_row ke setelah Bagian 1
        current_row = section_1_row + len(profile_data) + 1
        
        # ════════════════════════════════════════════════════════════════════════
        # BAGIAN 3 — MENU YANG DIREKOMENDASIKAN
        # ════════════════════════════════════════════════════════════════════════
        
        # Title Bagian 3
        ws[f'A{current_row}'] = "BAGIAN 3 — MENU YANG DIREKOMENDASIKAN"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Header tabel Bagian 3
        headers_3 = ['Waktu Makan', 'Menu/Makanan', 'Porsi (g/ml)', 'Kalori (kcal)', '']
        for col_idx, header in enumerate(headers_3, 1):
            try:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header  # type: ignore
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except (TypeError, ValueError):
                pass
        current_row += 1
        
        # Find kolom untuk Porsi dan Kalori di portion_result_df
        porsi_col = find_column_in_df(portion_result_df, [
            'recommended_portion_g', 'portion_g', 'adjusted_portion_g', 'gram'
        ])
        kalori_col = find_column_in_df(portion_result_df, [
            'adjusted_energy_kcal', 'energy_kcal_adjusted', 'final_energy_kcal', 'energy_kcal'
        ])
        nama_col = find_column_in_df(portion_result_df, [
            'food_name', 'name'
        ]) or find_column_in_df(selected_df, ['food_name', 'name'])
        slot_col = find_column_in_df(portion_result_df, [
            'slot', 'slot_name', 'meal_slot'
        ])
        
        print(f"[DEBUG export] Mapping: porsi_col={porsi_col}, kalori_col={kalori_col}, nama_col={nama_col}, slot_col={slot_col}")
        
        # Populate Bagian 3 dengan data
        for idx in range(len(selected_df)):
            # Get slot label
            if slot_col and idx < len(portion_result_df):
                slot_name = portion_result_df.iloc[idx].get(slot_col, SLOT_NAMES[idx] if idx < len(SLOT_NAMES) else f'Item {idx}')
            else:
                slot_name = SLOT_NAMES[idx] if idx < len(SLOT_NAMES) else f'Item {idx}'
            
            slot_label = SLOT_LABEL_MAP.get(slot_name, slot_name)
            
            # Get food name
            if idx < len(portion_result_df) and nama_col:
                food_name = portion_result_df.iloc[idx].get(nama_col, 'Unknown')
            else:
                food_name = selected_df.iloc[idx].get('food_name', 'Unknown')
            
            # Get porsi
            if idx < len(portion_result_df) and porsi_col:
                porsi = portion_result_df.iloc[idx].get(porsi_col, 0)
                porsi_str = f"{porsi:.1f}"
            else:
                porsi_str = "—"
            
            # Get kalori
            if idx < len(portion_result_df) and kalori_col:
                kalori = portion_result_df.iloc[idx].get(kalori_col, 0)
                kalori_str = f"{kalori:.1f}"
            else:
                kalori_str = "—"
            
            # Write row
            ws[f'A{current_row}'] = slot_label
            ws[f'B{current_row}'] = food_name
            ws[f'C{current_row}'] = porsi_str
            ws[f'D{current_row}'] = kalori_str
            
            for col_idx in range(1, 5):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border
                if col_idx in [3, 4]:  # Kolom Porsi dan Kalori
                    cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # ════════════════════════════════════════════════════════════════════════
        # BAGIAN 4 — DISTRIBUSI KALORI PER WAKTU MAKAN
        # ════════════════════════════════════════════════════════════════════════
        current_row += 2  # 2 baris kosong pemisah
        
        # Title Bagian 4
        ws[f'A{current_row}'] = "BAGIAN 4 — DISTRIBUSI KALORI PER WAKTU MAKAN"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Header tabel Bagian 4
        headers_4 = ['Waktu Makan', 'Kalori (kcal)', '% Aktual', '% Target', 'Keterangan']
        for col_idx, header in enumerate(headers_4, 1):
            try:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header  # type: ignore
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except (TypeError, ValueError):
                pass
        current_row += 1
        
        # Hitung kalori per meal
        meal_kalori_distribution = [
            ('Breakfast (Sarapan)', 0, 1, 2, 23.75),
            ('Lunch (Makan Siang)', 3, 4, 5, 33.75),
            ('Dinner (Makan Malam)', 6, 7, 8, 28.75),
            ('Snack', 9, None, None, 13.75),
        ]
        
        total_kalori_list = []
        for meal_name, idx1, idx2, idx3, target_pct in meal_kalori_distribution:
            meal_kalori = 0
            
            if kalori_col and idx1 < len(portion_result_df):
                meal_kalori += portion_result_df.iloc[idx1].get(kalori_col, 0)
            if idx2 is not None and kalori_col and idx2 < len(portion_result_df):
                meal_kalori += portion_result_df.iloc[idx2].get(kalori_col, 0)
            if idx3 is not None and kalori_col and idx3 < len(portion_result_df):
                meal_kalori += portion_result_df.iloc[idx3].get(kalori_col, 0)
            
            total_kalori_list.append(meal_kalori)
            
            # Hitung % Aktual: (kalori_meal / tdee * 100)
            pct_aktual = (meal_kalori / tdee * 100) if tdee > 0 else 0
            
            ws[f'A{current_row}'] = meal_name
            ws[f'B{current_row}'] = meal_kalori
            ws[f'C{current_row}'] = f"{pct_aktual:.1f}%"
            ws[f'D{current_row}'] = f"{target_pct:.2f}%"
            ws[f'E{current_row}'] = "Konstanta sistem"
            
            for col_idx in range(1, 6):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border
                if col_idx == 2:  # Kolom Kalori
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '0.0'
                elif col_idx in [3, 4]:  # Kolom % Aktual dan % Target
                    cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # ════════════════════════════════════════════════════════════════════════
        # BAGIAN 5 — HARD CONSTRAINT
        # ════════════════════════════════════════════════════════════════════════
        current_row += 2  # 2 baris kosong pemisah
        
        # Title Bagian 5
        ws[f'A{current_row}'] = "BAGIAN 5 — HARD CONSTRAINT (DISEASE-BASED)"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Header tabel Bagian 5
        headers_5 = ['Nutrisi', 'Min – Max', 'Aktual', 'Keterpenuhan (%)', '']
        for col_idx, header in enumerate(headers_5, 1):
            try:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header  # type: ignore
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except (TypeError, ValueError):
                pass
        current_row += 1
        
        # Populate Bagian 5
        for nutrient_key, constraint in sorted(guidelines['hard'].items()):
            min_val = constraint.get('min', 0)
            max_val = constraint.get('max', float('inf'))
            max_str = "∞" if max_val == float('inf') else f"{max_val:.1f}"
            range_str = f"{min_val:.1f} – {max_str}"
            
            # Lookup aktual dari total_nutrition
            actual = lookup_nutrition(total_nutrition, nutrient_key, HARD_NUTRIENT_KEY_MAP)
            
            if actual is not None:
                pct = calculate_fulfillment_percentage(actual, min_val, max_val)
                pct_str = f"{pct:.1f}%"
                actual_str = f"{actual:.1f}"
            else:
                pct_str = "N/A"
                actual_str = "N/A"
            
            ws[f'A{current_row}'] = nutrient_key.replace('_', ' ').title()
            ws[f'B{current_row}'] = range_str
            ws[f'C{current_row}'] = actual_str
            ws[f'D{current_row}'] = pct_str
            
            for col_idx in range(1, 5):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border
                if col_idx in [2, 3, 4]:  # Kolom range, actual, percentage
                    cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # ════════════════════════════════════════════════════════════════════════
        # BAGIAN 6 — SOFT CONSTRAINT
        # ════════════════════════════════════════════════════════════════════════
        current_row += 2  # 2 baris kosong pemisah
        
        # Title Bagian 6
        ws[f'A{current_row}'] = "BAGIAN 6 — SOFT CONSTRAINT (DRI-BASED)"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Header tabel Bagian 6
        headers_6 = ['Nutrisi', 'Target (DRI)', 'Aktual', 'Keterpenuhan (%)', 'Status']
        for col_idx, header in enumerate(headers_6, 1):
            try:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header  # type: ignore
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            except (TypeError, ValueError):
                pass
        current_row += 1
        
        # Populate Bagian 6
        for nutrient_key, constraint in sorted(guidelines['soft'].items()):
            min_val = constraint.get('min', 0)
            max_val = constraint.get('max', float('inf'))
            max_str = "∞" if max_val == float('inf') else f"{max_val:.1f}"
            target_str = f"{min_val:.1f} – {max_str}"
            
            actual = lookup_nutrition(total_nutrition, nutrient_key, SOFT_NUTRIENT_KEY_MAP)
            
            if actual is not None:
                pct = calculate_fulfillment_percentage(actual, min_val, max_val)
                status_text, _, _ = get_status_category(pct)
                actual_str = f"{actual:.2f}"
                pct_str = f"{pct:.1f}%"
            else:
                status_text = "—"
                actual_str = "N/A"
                pct_str = "N/A"
            
            ws[f'A{current_row}'] = nutrient_key.replace('_', ' ').title()
            ws[f'B{current_row}'] = target_str
            ws[f'C{current_row}'] = actual_str
            ws[f'D{current_row}'] = pct_str
            ws[f'E{current_row}'] = status_text
            
            for col_idx in range(1, 6):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border
                if col_idx in [2, 3, 4]:  # Kolom target, actual, keterpenuhan
                    cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        # ════════════════════════════════════════════════════════════════════════
        # SAVE FILE
        # ════════════════════════════════════════════════════════════════════════
        wb.save(filename)
        
        # Print success message dengan path absolut
        abs_path = os.path.abspath(filename)
        print(f"\n✓ EXPORT SUKSES!")
        print(f"  File: {abs_path}")
        print(f"  Sheet: {sheet_name}")
        
        return True
    
    except Exception as e:
        print(f"\n✗ EXPORT GAGAL!")
        print(f"  Error: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_ga_with_nutrition_service():
    """
    Main flow: User input → NutritionService → GA → Output
    
    Steps:
        1. Get user input
        2. Call NutritionService.calculate_nutrition_needs()
        3. Extract food_df dan guidelines
        4. Run GA
        5. Display hasil
    """
    
    try:
        # STEP 1: Get user input
        print("\nSTEP 1: Get user input...")
        
        if USE_INTERACTIVE_INPUT:
            # Use interactive input handler from modules
            user_input = get_user_input()
        else:
            # Use default values (fallback)
            user_input = get_simple_user_input(interactive=False)
        
        print("\n✓ User input received")
        print(f"  Gender: {user_input['gender']}")
        print(f"  Age: {user_input['age']}, Weight: {user_input['weight']}kg, Height: {user_input['height']}cm")
        print(f"  Activity Factor: {user_input['activity_factor']}")
        print(f"  Diseases: {user_input['disease']}")
        print(f"  Food Preferences: {user_input['food_preferences']}")
        
        # STEP 2: Calculate nutrition requirements using NutritionService
        print("\n" + "="*70)
        print("STEP 2: Calculate nutrition requirements...")
        print("="*70)
        
        service = NutritionService()
        nutrition_result = service.calculate_nutrition_needs(user_input)
        
        # Check success
        if not nutrition_result['success']:
            print(f"✗ FAILED: {nutrition_result.get('error', 'Unknown error')}")
            return
        
        print("✓ Nutrition calculation successful")
        
        # STEP 3: Extract data dari nutrition_result
        print("\nSTEP 3: Extract data from NutritionService...")
        
        food_df = nutrition_result['food_data']['dataframe']
        guidelines_all = nutrition_result['guidelines']['nutrients']
        tdee = nutrition_result['energy']['tdee']
        
        # ════════════════════════════════════════════════════════════════════════
        # NEW: SPLIT GUIDELINES MENJADI HARD DAN SOFT CONSTRAINTS
        # ════════════════════════════════════════════════════════════════════════
        # HARD constraint: tipe="range" atau "max" dari guideline.csv (HIGH PRIORITY)
        # SOFT constraint: DRI micronutrient (FLEXIBLE)
        
        # Split guidelines berdasarkan hard_soft_type
        guidelines = {
            'hard': {k: v for k, v in guidelines_all.items() if v.get('hard_soft_type') == 'HARD'},
            'soft': {k: v for k, v in guidelines_all.items() if v.get('hard_soft_type') != 'HARD'}
        }
        
        print(f"✓ Data extracted:")
        print(f"  - Food items available: {len(food_df)}")
        print(f"  - HARD constraints: {len(guidelines['hard'])} nutrients")
        print(f"  - SOFT constraints: {len(guidelines['soft'])} nutrients")
        print(f"  - User TDEE: {tdee:.0f} kcal/day")
        
        # Display some info dari NutritionService
        print(f"\n📊 User Profile:")
        anthro = nutrition_result['anthropometrics']
        print(f"  - BMI: {anthro['bmi']:.1f} ({anthro['bmi_category']})")
        print(f"  - BBI: {anthro['bbi']:.1f} kg")
        energy = nutrition_result['energy']
        print(f"  - BMR: {energy['bmr']:.0f} kcal/day")
        print(f"  - TDEE: {energy['tdee']:.0f} kcal/day")
        
        # Display HARD vs SOFT constraints
        print(f"\n🎯 HARD Constraints (Disease-based - HIGH PRIORITY):")
        for nutrient in sorted(guidelines['hard'].keys()):
            constraint = guidelines['hard'][nutrient]
            min_val = constraint.get('min', 0)
            max_val = constraint.get('max', float('inf'))
            unit = constraint.get('unit', 'unit')
            print(f"  - {nutrient:20s}: {min_val:8.1f} - {max_val:8.1f} {unit}")
        
        print(f"\n🎯 SOFT Constraints (DRI-based - FLEXIBLE):")
        for nutrient, constraint in sorted(guidelines['soft'].items()):
            min_val = constraint.get('min', 0)
            max_val = constraint.get('max', float('inf'))
            max_str = f"{max_val:10.2f}" if max_val != float('inf') else "       inf"
            print(f"  - {nutrient:20} : {min_val:10.2f} - {max_str}")
        
        # STEP 5: Run GA
        print("="*70)
        print("STEP 5: Run Genetic Algorithm...")
        print("="*70)
        
        best_solution, top_solutions = run_ga(
            food_df=food_df,
            guidelines=guidelines,
            tdee=tdee,
            generations=100,
            pop_size=50,
            elite_ratio=0.15,
            mutation_rate=0.35,
            verbose=False  # Changed to False for cleaner output
        )
        print("✓ GA optimization complete")
        
        # STEP 5.5: LOCAL SEARCH - Fine-tuning untuk meningkatkan solusi
        print("\n" + "="*70)
        print("STEP 5.5: Local Search - Fine-tuning GA Result...")
        print("="*70)
        
        best_solution = local_search(
            solution=best_solution,
            food_df=food_df,
            guidelines=guidelines,
            tdee=tdee,
            iterations=50,
            verbose=True  # Show improvements
        )
        print("✓ Local search optimization complete")
        
        # STEP 6: Display hasil
        print("\n" + "="*70)
        print("STEP 6: OPTIMAL MEAL PLAN - GA RESULT")
        print("="*70)
        
        display_solution(best_solution, guidelines)
        display_fitness_details(best_solution, guidelines)
        
        # Tambahkan rekomendasi khusus berdasarkan disease
        if any('dm2' in disease.lower() for disease in user_input['disease']):
            print("\n" + "⚠️  REKOMENDASI TAMBAHAN:")
            print("    Konsumsi buah & sayur minimal 400g/hari untuk mendukung")
            print("    kontrol gula darah pada pasien Diabetes Mellitus Tipe 2.")
        
        # STEP 7: Generate meal options dari top_solutions (berbagai kombinasi)
        print("\n" + "="*70)
        print("STEP 7: Generate 2-3 varied menu options per slot...")
        print("="*70)
        
        # Gabungkan best_solution (setelah LS) sebagai opsi pertama
        # bersama top_solutions dari GA sebagai variasi
        top_solutions_with_best = [best_solution] + top_solutions
        slot_options = generate_meal_options(
            food_df,
            top_solutions_with_best,
            max_options_per_slot=3,
            food_preferences=user_input['food_preferences']
        )
        print("✓ Menu options generated: 10 slots ready")
        
        # ============================================================================
        # STEP 8: USER SELECTION - Interactive menu selection
        # ============================================================================
        print("\n" + "="*70)
        print("STEP 8: USER SELECTION - Choose your menu")
        print("="*70)
        
        selected_meal = []
        
        # Mapping slot name ke meal type untuk better display
        meal_display_map = {
            'breakfast_main': ('BREAKFAST', 'MAIN'),
            'breakfast_side': ('BREAKFAST', 'SIDE'),
            'breakfast_drink': ('BREAKFAST', 'DRINK'),
            'lunch_main': ('LUNCH', 'MAIN'),
            'lunch_side': ('LUNCH', 'SIDE'),
            'lunch_drink': ('LUNCH', 'DRINK'),
            'dinner_main': ('DINNER', 'MAIN'),
            'dinner_side': ('DINNER', 'SIDE'),
            'dinner_drink': ('DINNER', 'DRINK'),
            'snack': ('SNACK', 'ITEM')
        }
        
        # Loop setiap slot dan minta user untuk memilih
        for slot_idx, slot_name in enumerate(SLOT_NAMES):
            options = slot_options.get(slot_name, [])
            
            if not options:
                print(f"\n⚠️  {slot_name}: Tidak ada opsi tersedia")
                continue
            
            meal_type, item_type = meal_display_map.get(slot_name, (slot_name, 'Item'))
            
            print(f"\n{'─' * 70}")
            print(f"{meal_type} - {item_type} (Slot {slot_idx})")
            print(f"{'─' * 70}")
            
            # Display 3 options
            for i, option in enumerate(options, 1):
                food_name = option.get('food_name', 'Unknown')
                energy = option.get('energy_kcal', 0)
                protein = option.get('protein_g', 0)
                print(f"{i}. {food_name:30} | Energy: {energy:6.1f} kcal | Protein: {protein:5.1f}g")
            
            # Get user choice
            while True:
                try:
                    choice_str = input(f"\nPilih opsi (1-{len(options)}) [default=1]: ").strip()
                    
                    # Default to 1 jika user tekan Enter
                    if choice_str == "":
                        choice = 0
                    else:
                        choice = int(choice_str) - 1
                    
                    if 0 <= choice < len(options):
                        selected_item = options[choice].copy()
                        selected_meal.append(selected_item)
                        print(f"✓ {options[choice].get('food_name', 'Unknown')} dipilih")
                        break
                    else:
                        print(f"✗ Pilih antara 1-{len(options)}")
                except ValueError:
                    print("✗ Input harus berupa angka")
        
        # Convert selected meals ke DataFrame
        print("\n" + "="*70)
        print("Memproses pilihan user...")
        print("="*70)
        
        if len(selected_meal) == CHROMOSOME_SIZE:
            selected_df = pd.DataFrame(selected_meal).reset_index(drop=True)
            print(f"✓ {len(selected_df)} items dipilih dari {CHROMOSOME_SIZE} slots")
            
            # Calculate total nutrition dari selected meals
            selected_nutrition = calculate_total_nutrition(selected_df)
            
            # STEP 9: Display final nutrition comparison (simplified)
            print("\n" + "="*70)
            print("STEP 9: NUTRITION ANALYSIS - Your Selected Menu")
            print("="*70)
            
            print("\n📋 YOUR SELECTED MENU (10 Items):")
            print("─" * 70)
            
            # Display meals by meal type
            for meal in ['breakfast', 'lunch', 'dinner', 'snack']:
                indices = MEAL_INDICES[meal]
                meal_items = [selected_df.iloc[i].get('food_name', f'Item {i}') 
                             for i in indices if i < len(selected_df)]
                print(f"\n{meal.upper():12}: {' | '.join(meal_items)}")
            
            # Calculate total nutrition dari selected meals
            selected_nutrition = calculate_total_nutrition(selected_df)
            
            print("\n" + "─" * 70)
            print("📊 NUTRITION ANALYSIS:")
            print("─" * 70)
            
            # Display detailed nutrition analysis for selected menu
            display_nutrition_analysis_table(selected_df, guidelines)
            
            # ════════════════════════════════════════════════════════════════════════
            # STEP 10: PORTION SIZING - Calculate portion sizes dynamically (MEAL-BASED + DEFICIT-AWARE)
            # ════════════════════════════════════════════════════════════════════════
            portion_result_df = calculate_portion_sizes_dynamic(selected_df, tdee, guidelines)
            display_portion_summary_dynamic(portion_result_df, guidelines, tdee)
            
            # ════════════════════════════════════════════════════════════════════════
            # STEP 11: EXPORT HASIL KE EXCEL
            # ════════════════════════════════════════════════════════════════════════
            excel_filename = input("\nMasukkan nama file Excel output (tanpa .xlsx): ").strip()
            if not excel_filename:
                excel_filename = "hasil_meal_plan"
            excel_filename += ".xlsx"
            
            # Call export_to_excel
            export_to_excel(
                filename=excel_filename,
                user_input=user_input,
                nutrition_result=nutrition_result,
                guidelines_all=guidelines_all,
                selected_df=selected_df,
                portion_result_df=portion_result_df,
                guidelines=guidelines,
                tdee=tdee
            )
            
            print("\n✓ MEAL PLANNING SYSTEM - COMPLETE")
            print("="*70 + "\n")
        else:
            print(f"✗ Error: Hanya {len(selected_meal)} dari {CHROMOSOME_SIZE} items yang dipilih")
            print("\n✓ MEAL PLANNING SYSTEM - COMPLETE")
            print("="*70 + "\n")
    
    except ValueError as e:
        print(f"\n✗ VALUE ERROR: {e}")
        import traceback
        traceback.print_exc()
    except KeyError as e:
        print(f"\n✗ KEY ERROR: {e}")
        import traceback
        traceback.print_exc()
    except Exception as e:
        print(f"\n✗ UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    test_ga_with_nutrition_service()
